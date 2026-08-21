# =================== IMPORTS ====================
import pdfplumber
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment # fonte e alinhamento
import logging
from datetime import datetime
from time import perf_counter # cronômetro de alta precisão pra medir duração
from pathlib import Path
from collections import Counter
# ================================================

Path('Logs').mkdir(exist_ok=True) # pasta onde se encontrará o log

caminho_logs = Path('Logs') / 'Faturas.log' # Definindo o caminho certo para onde o log será guardado

# Configuração básica do logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler(caminho_logs, encoding='utf-8'), # Salvando histórico em arquivo de log
                        logging.StreamHandler() # Exibindo Histórico ao vivo no terminal.
                    ])

# Separador de execuções
logging.info('=' * 50)
logging.info(f'EXECUÇÃO: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}') # strftime -> datetime vira str
logging.info('=' * 50)

# Marca o início da execução, pra cronometrar o run inteiro
inicio = perf_counter()

# Dicionário de referência para padronização de datas
meses = {
    'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
    'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
    'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'
}


pasta_faturas = Path('Faturas_teste') # pasta onde se encontrará todos os arquivos de faturas a serem lidos

Path('Saidas').mkdir(exist_ok=True) # pasta de outputs (xlsx) para o usuário 

caminho_saida = Path('Saidas') / 'Faturas_teste.xlsx' # Definindo o caminho certo para onde o xlsx será guardado

def identificador(arquivo):
    with pdfplumber.open(arquivo) as pdf:
        texto = pdf.pages[0].extract_text() or ''
        if 'vigente' in texto:
            return nubank, 'Nubank'
        elif 'Sicredi' in texto:
            return sicredi, 'Sicredi'
        elif 'OUROCARD' in texto:
            return banco_do_brasil, 'Banco do Brasil'
        elif 'A sua fatura chegou' in texto:
            return inter, 'Inter'
        else:
            return None, None

# ======================================= PARSERS =======================================

# Nubank
def nubank(arquivo):
    # Leitura da fatura em PDF
    def ler_nubank():
        transacoes = []
        logging.info('Iniciando Leitura: Fatura Nubank.')
        lixo = ('IOF', 'Juros', 'rotativo', 'Total de compras')  # encargos e a linha de total
        try:
            with pdfplumber.open(arquivo) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text().splitlines() # Separando texto em linhas
                    for linha in texto:
                        tem_data = re.search(r'\d{2}\s[A-Z]{3}', linha) # ex: 01 ABR
                        tem_valor = re.search(r'R\$\s*[\d.]+,\d{2}', linha) # ex: R$ 100,00
                        eh_negativo = re.search(r'[-\u2212]\s*R\$', linha) # pagamento/crédito
                        eh_lixo = any(l in linha for l in lixo) # encargo ou total
                        if tem_data and tem_valor and not eh_negativo and not eh_lixo:
                            transacoes.append(linha)
            logging.info('Leitura concluída')
        except FileNotFoundError as e:
            logging.error(f'Arquivo não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return transacoes

    # Extração das transações usando Regex; passando para Dict e construindo lista de Dict
    def extrair_dados(leitura):
        dados_transacoes = []
        logging.info('Iniciando extração de dados: Fatura Nubank.')
        try:
            ano_emissao, mes_emissao = data_emissao_pdf(arquivo) # Ano e mês da emissão, lidos do metadado do PDF (uma vez, fora do loop)
            for compras in leitura:
                # Formato de data original extraída do arquivo
                data_compra = re.findall(r'\d{2}\s[A-Z]{3}', compras)[0] # Ex. 12 ABR

                # Separando
                dia, mes_abreviado = data_compra.split()
                mes_num = meses[mes_abreviado]

                # Padronizando a data
                data_padronizada = montar_data(dia, mes_num, ano_emissao, mes_emissao)

                descricao_completa = re.sub(r'\d{2}\s[A-Z]{3}|••••\s\d{4}|Parcela\s\d+/\d+|R\$\s*[\d,.]+', '', compras).replace('-', '').strip()
                descricao_limpa = descricao_completa.split('\n')[0].strip()
                parcela_encontrada = re.findall(r'\d+/\d+', compras)
                parcela = parcela_encontrada[0] if parcela_encontrada else '1/1'
                valor_compras = re.findall(r'R\$\s*[\d,.]+', compras)[0]
                # Dict de compras
                dict_compras = {'Data': data_padronizada,
                                'Descrição': descricao_limpa,
                                'Parcela': parcela,
                                'Valor': valor_compras}
                dados_transacoes.append(dict_compras) # Lista de Dict de compras
            logging.info('Extração de dados concluída.')
        except IndexError as e:
            logging.error(f'Padrão não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return dados_transacoes

    resultado_leitura = ler_nubank()
    dados_extraidos = extrair_dados(resultado_leitura)

    texto = texto_completo(arquivo_fatura=arquivo)
    
    saldo_f = extrair_valores(texto, 'Saldo financiado')
    juros = extrair_valores(texto, 'Juros de financiamento')
    iof_financiamento = extrair_valores(texto, 'IOF de financiamento')
    iof_internacional = extrair_valores(texto, 'IOF de compras internacionais')

    if saldo_f:
        dados_extraidos.append({'Data': None,
                                'Descrição': 'Saldo financiado',
                                'Parcela': '1/1',
                                'Valor': f'R$ {saldo_f}'})
    if juros:
        dados_extraidos.append({'Data': None,
                                'Descrição': 'Juros de financiamento',
                                'Parcela': '1/1',
                                'Valor': f'R$ {juros}'})
    if iof_financiamento:
        dados_extraidos.append({'Data': None,
                                'Descrição': 'IOF de financiamento',
                                'Parcela': '1/1',
                                'Valor': f'R$ {iof_financiamento}'})
    if iof_internacional:
        dados_extraidos.append({'Data': None,
                                'Descrição': 'IOF de compras internacionais',
                                'Parcela': '1/1',
                                'Valor': f'R$ {iof_internacional}'})

    total_fatura = extrair_valores(texto, 'Total a pagar R$', inicio=True)

    if total_fatura is not None:
        total_fatura = para_float(total_fatura)

    return dados_extraidos, total_fatura

# Inter
def inter(arquivo):
    # Leitura do arquivo PDF
    def ler_inter():
        transacoes = []
        logging.info('Iniciando Leitura: Fatura Inter.')
        try:
            with pdfplumber.open(arquivo) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text().splitlines() # Separando texto em linhas
                    for linha in texto:
                        if re.search(r'\d{2}\sde\s[a-z]+\.\s\d{4}', linha) and '+' not in linha:
                            transacoes.append(linha)
            logging.info('Leitura concluída.')
        except FileNotFoundError as e:
            logging.error(f'Arquivo não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return transacoes

    # Extração das transações usando Regex; passando para Dict e construindo lista de Dict
    def extrair_dados(leitura):
        dados_transacoes = []
        logging.info('Iniciando extração de dados: Fatura Inter.')
        try:
            for compras in leitura:
                # Formato de data original extraída do arquivo
                data_compra = re.findall(r'\d{2}\sde\s[a-z]+\.\s\d{4}', compras)[0] # ex: 01 de jan. 2025

                # Usando regex de grupos para separar - forma alternativa ao split()
                match = re.search(r'(\d{2})\sde\s([a-z]+)\.\s(\d{4})', data_compra)
                dia = match.group(1)
                mes_abrevidado = match.group(2)
                ano = match.group(3)

                # Convertendo mês para upper - padrão do dicionário
                mes_num = meses[mes_abrevidado.upper()]

                # Padronizando a data
                data_padronizada = datetime.strptime(f'{dia}/{mes_num}/{ano}', '%d/%m/%Y').date()

                descricao = re.sub(r'\d{2}\sde\s[a-z]+\.\s\d{4}|\(Parcela.*?\)|R\$\s*[\d,.]+', '', compras).replace('-', '').strip()
                parcela_encontrada = re.findall(r'\d+\sde\s\d+', compras)
                parcela = parcela_encontrada[0] if parcela_encontrada else '1/1'
                valor_compras = re.findall(r'R\$\s*[\d,.]+', compras)[0]
                # Dict de compras
                dict_compras = {'Data': data_padronizada,
                                'Descrição': descricao,
                                'Parcela': parcela.replace('de', '/').replace(' ', ''),
                                'Valor': valor_compras}
                dados_transacoes.append(dict_compras) # Lista de Dict de compras
            logging.info('Extração de dados concluída: Fatura Inter.')
        except IndexError as e:
            logging.error(f'Padrão não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return dados_transacoes

    resultado_leitura = ler_inter()
    dados_extraidos = extrair_dados(resultado_leitura)

    texto = texto_completo(arquivo_fatura=arquivo)
    total_fatura = extrair_valores(texto, 'VALOR DO DOCUMENTO')
    if total_fatura is not None:
        total_fatura = para_float(total_fatura)

    return dados_extraidos, total_fatura

# Sicredi
def sicredi(arquivo):
    # Leitura da fatura em PDF
    def ler_sicredi():
        transacoes = []
        logging.info('Iniciando leitura: Fatura Sicredi.')
        try:
            with pdfplumber.open(arquivo) as pdf:
                for pagina in pdf.pages:
                    tabela = pagina.extract_table() # Sicredi prioriza o uso de tabelas
                    if tabela:
                        for linha in tabela:
                            if linha[0] and re.search(r'\d{2}/[a-z]{3}\s\d{2}:\d{2}', linha[0]) and '-R$' not in linha[0]:
                                transacoes.append(linha)
            logging.info('Leitura concluída.')
        except FileNotFoundError as e:
            logging.error(f'Arquivo não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado {e}')
        return transacoes

    # Extração das transações usando Regex; passando para Dict e construindo lista de Dict
    def extrair_dados(leitura):
        dados_transacoes = []
        logging.info('Iniciando extração de dados: Fatura Sicredi.')
        try:
            ano_emissao, mes_emissao = data_emissao_pdf(arquivo) # Ano e mês da emissão, lidos do metadado do PDF (uma vez, fora do loop)
            for compras in leitura:
                # Usando regex de grupos para separar - forma alternativa ao split()
                match = re.search(r'(\d{2})/([a-z]{3})', compras[0])
                dia = match.group(1)
                mes_abreviado = match.group(2)

                # Convertendo mês para upper - padrão do dicionário
                mes_num = meses[mes_abreviado.upper()]

                # Padronizando a data
                data_padronizada = montar_data(dia, mes_num, ano_emissao, mes_emissao)

                partes = compras[0].splitlines()
                # divide a string nos \n -> lista de linhas
                if len(partes) > 1:  # tem \n -> é internacional
                    if partes[0].startswith('PYG'):  # começa com PYG?
                        # descrição está na segunda linha
                        linha_limpa = re.sub(r'\d{2}/[a-z]{3}\s\d{2}:\d{2}', '', partes[1])
                        descricao = linha_limpa.split('R$')[0].strip()
                    else:
                        # descrição está na primeira linha antes do PYG
                        descricao = partes[0].split('PYG')[0].strip()
                else:  # sem \n -> é nacional
                    linha_limpa = re.sub(r'\d{2}/[a-z]{3}\s\d{2}:\d{2}', '', partes[0])
                    descricao = linha_limpa.split('R$')[0].strip()
                parcela_encontrada = re.findall(r'\d+/\d+', compras[0])
                parcela = parcela_encontrada[0] if parcela_encontrada else '1/1'
                valor = re.findall(r'R\$\s*[\d,.]+', compras[0])[-1]
                # Dicionario de transações
                dict_compras = {'Data': data_padronizada,
                                'Descrição': descricao,
                                'Parcela': parcela,
                                'Valor': valor}
                dados_transacoes.append(dict_compras)
            logging.info('Extração de dados concluida: Fatura Sicredi.')
        except IndexError as e:
            logging.error(f'Padrão não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return dados_transacoes

    resultado_leitura = ler_sicredi()
    dados_extraidos = extrair_dados(resultado_leitura)

    texto = texto_completo(arquivo_fatura=arquivo)
    encargo = extrair_valores(texto, 'Subtotal Encargos')

    if encargo:
        dados_extraidos.append({'Data': None,
                                'Descrição': 'Encargos',
                                'Parcela': '1/1',
                                'Valor': f'R$ {encargo}'})

    total_fatura = extrair_valores(texto, 'Total desta Fatura')
    if total_fatura is not None:
        total_fatura = para_float(total_fatura)

    return dados_extraidos, total_fatura

# Banco do Brasil
def banco_do_brasil(arquivo):
    # Leitura da fatura em PDF
    def ler_bb():
        transacoes = []
        logging.info('Iniciando leitura: Fatura Banco do Brasil.')
        try:
            with pdfplumber.open(arquivo) as pdf:
                for pagina in pdf.pages[1:]: # pula a página 0: ela tem uma linha em formato de data que o regex capturaria como transação falsa
                    texto = pagina.extract_text().splitlines() # Dividindo o texto em linhas
                    for linha in texto:
                        if re.search(r'^\d{2}/\d{2}', linha) and not re.search(r'R\$\s*-[\d,.]+', linha):
                            transacoes.append(linha)
            logging.info('Leitura concluída.')
        except FileNotFoundError as e:
            logging.error(f'Arquivo não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return transacoes

    # Extração das transações usando Regex; passando para Dict e construindo lista de Dict
    def extrair_dados(leitura):
        dados_transacoes = []
        logging.info('Iniciando extração de dados: Fatura Banco do Brasil.')
        try:
            ano_emissao, mes_emissao = data_emissao_pdf(arquivo) # Ano e mês da emissão, lidos do metadado do PDF (uma vez, fora do loop)
            for compras in leitura:
                data_compra = re.findall(r'^\d{2}/\d{2}', compras)[0] # data original do PDF
                dia, mes = data_compra.split('/')
                data_padronizada = montar_data(dia, mes, ano_emissao, mes_emissao) # Data padronizada
                descricao = re.sub(r'^\d{2}/\d{2}|R\$\s*[\d,.]+|PARC\s\d+/\d+', '', compras).strip()
                linha_sem_data = re.sub(r'^\d{2}/\d{2}\s', '', compras)
                parcela_encontrada = re.findall(r'PARC\s(\d+/\d+)', linha_sem_data)
                parcela = parcela_encontrada[0] if parcela_encontrada else '1/1'
                valores = re.findall(r'R\$\s*[\d,.]+', compras)
                valor = valores[0] if valores else None
                dict_compras = {'Data':data_padronizada,
                                'Descrição': descricao,
                                'Parcela': parcela,
                                'Valor': valor}
                dados_transacoes.append(dict_compras)
            logging.info('Extração de dados concluída: Fatura Banco do Brasil.')
        except IndexError as e:
            logging.error(f'Padrão não encontrado: {e}')
        except Exception as e:
            logging.exception(f'Erro inesperado: {e}')
        return  dados_transacoes

    resultado_leitura = ler_bb()
    dados_extraidos = extrair_dados(resultado_leitura)

    texto = texto_completo(arquivo_fatura=arquivo)
    total_fatura = extrair_valores(texto, 'Total da Fatura')
    if total_fatura is not None:
        total_fatura = para_float(total_fatura)

    return dados_extraidos, total_fatura

# ======================================= HELPERS =======================================

# Lê ano e mês da emissão da fatura no metadado do PDF (CreationDate); usa o relógio se não houver
def data_emissao_pdf(arquivo_fatura): 
    with pdfplumber.open(arquivo_fatura) as pdf:
        data_criacao_pdf = pdf.metadata.get('CreationDate')
        if data_criacao_pdf:
            return int(data_criacao_pdf[2:6]), int(data_criacao_pdf[6:8])
        agora = datetime.now()
        return agora.year, agora.month

def data_lote_emissao_pdf(pdfs_validos):
    datas = []
    for pdf in pdfs_validos:
        emissao = data_emissao_pdf(pdf)
        datas.append((pdf, emissao))
    return datas

def mes_saida(datas):
    emissoes = [emissao for pdf, emissao in datas]
    contador = Counter(emissoes).most_common(1)
    vencedor = contador[0][0]
    for pdf, emissao in datas:
        if emissao != vencedor:
            logging.warning(f'{pdf.name}: emissão {emissao} diverge do lote {vencedor}')
    return vencedor

# Monta a data da compra corrigindo o ano na virada: compra de mês posterior à emissão é do ano anterior
def montar_data(dia, mes_compra, ano_emissao, mes_emissao):
    if int(mes_compra) > int(mes_emissao):
        ano = ano_emissao - 1
    else:
        ano = ano_emissao
    return datetime.strptime(f'{dia}/{mes_compra}/{ano}', '%d/%m/%Y').date()


# Extrai o texto de todas as páginas do PDF (documento inteiro)
def texto_completo(arquivo_fatura):
    texto = ''
    with pdfplumber.open(arquivo_fatura) as pdf:
        for pagina in pdf.pages:
            texto += (pagina.extract_text() or '') + '\n' 
        return texto

# Extraindo valores
def extrair_valores(texto, ancora, inicio=False): # Ancora seria o texto que preciso. Ex.: "Encargos"
    for linha in texto.splitlines():
        if inicio:
            achou = linha.strip().startswith(ancora)
        else:
            achou = ancora in linha
        if achou:
            valores = re.findall(r'[\d.]+,\d{2}', linha) # Pegando valor que está na ancora que passei.
            if valores:
                return valores[-1]
    return None

# Converter para float
def para_float(valor):
    return float(valor.replace('R$', '').replace('.', '').replace(',', '.').strip())

# Confere a soma extraída pelo Python contra o valor real do arquivo
def conferir(soma_extraida, total_fatura, banco):

    if total_fatura is None:
        msg = 'Impossível conferir: Total não encontrado'
        logging.warning(f'{banco}: {msg}')
        return msg

    diferença = total_fatura - soma_extraida

    if abs(diferença) < 0.10:
        msg = (f'Tudo certo')
    elif soma_extraida < total_fatura:
        msg = (f'Falta R$ {diferença:.2f}')
    else:
        msg = (f'sobra R$ {-diferença:.2f}')

    logging.info(f'{banco}: {msg}')
    return msg

# Monta o DataFrame e acrescenta as colunas de preenchimento manual
def preparar_df(transacoes):
    try:

        logging.info('Criando DataFrame')

        df = pd.DataFrame(transacoes)
        # Coluna Titular
        df.insert(1, 'Titular', None) # Preenchimento manual
        # Capitaliza a descrição (1ª letra maiúscula) — padroniza o visual de todos os bancos
        df['Descrição'] = df['Descrição'].str.capitalize()
        # Conversão da coluna 'valor' de str para float para permitir operações numéricas.
        df['Valor'] = df['Valor'].apply(para_float)
        # Coluna valor pago
        df['Valor pago'] = None # Preenchimento manual
        # Coluna valor restante
        df['Valor restante'] = None # Preenchimento manual

        logging.info('DataFrame criado.')

        return df
    except KeyError as e:
        logging.error(f'Coluna não encontrada: {e}')
    except ValueError as e:
        logging.error(f'Erro na conversão de valor: {e}')
    except TypeError as e:
        logging.error(f'Tipo inválido: {e}')
    except Exception as e:
        logging.error(e)

# Exportar arquivo
def exportar_xlsx(fatura, workbook):
    try:
        logging.info('Exportando arquivo .xlsx')
        with pd.ExcelWriter(workbook) as writer:
            for df, feedback, sheet in fatura:
                df.to_excel(writer, sheet_name=sheet, index=False)
                writer.sheets[sheet]['K1'] = feedback
        logging.info('Exportação concluída.')
    except PermissionError as e:
        logging.error(f'Feche o arquivo Excel antes de continuar: {e}')
    except ValueError as e:
        logging.error(f'Nome da sheet inválido. Caractéres especiais ou nome muito longo: {e}')
    except Exception as e:
        logging.error(e)

# Formatação
def manipular_xlsx(workbook):
    # Carrega o workbook (pasta de trabalho)
    try:
        logging.info('Iniciando formatação.')
        wb = openpyxl.load_workbook(workbook)

        # Adiciona a fórmula de cálculo na coluna "Valor restante"
        def adicionar_calculo():
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in range(2, ws.max_row + 1):
                    ws[f'G{row}'] = f'=E{row}-F{row}'

        # Formatar cabeçalho
        def formatar_cabecalho():
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for cell in ws[1]:
                    cell.font = Font(size=13,
                                     name='Arial',
                                     bold=True)
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center')

        # Fonte e alinhamento das linhas de dados
        def fonte():
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(size=12,
                                         name='Arial')
                        cell.alignment = Alignment(horizontal='left',
                                                   vertical='center')
                # centralizando somente a coluna de parcelas
                for row in ws.iter_cols(min_col=4, max_col=4):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center',
                                                   vertical='center')
                 
        # Formatando campos de valores e passando para forma monetária R$
        def campos_monetarios():
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for column in ws.iter_cols(min_col=5, max_col=8):
                    for cell in column:
                        cell.number_format = 'R$ #,##0.00'

        # Formatando campos de datas e passando para formato visual no Brasil - DD/MM/YYYY -> dia/mês/ano
        def campos_data():
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for column in ws.iter_cols(min_col=1, max_col=1, min_row=2):
                    for cell in column:
                        cell.number_format = 'DD/MM/YYYY'

        # Quantos caracteres a célula ocupa EXIBIDA — dinheiro e data formatados são mais longos que o valor cru
        def tamanho_celula(cell):
            if cell.value is None:
                return 0
            formato = cell.number_format
            if 'R$' in formato and isinstance(cell.value, (int, float)):
                return len(f'R$ {cell.value:,.2f}')
            if 'YYYY' in formato:
                return 12

            return len(str(cell.value))

        # Ajusta a largura de cada coluna ao maior conteúdo (openpyxl não tem autofit nativo; o +3 compensa a fonte)
        def auto_fit():
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for coluna in ws.columns:
                    maior = max((tamanho_celula(cell) for cell in coluna), default=0) # defaul para coluna vazia
                    letra = coluna[0].column_letter
                    ws.column_dimensions[letra].width = maior + 3

        # Chamadas Internas
        adicionar_calculo()
        formatar_cabecalho()
        fonte()
        campos_monetarios()
        campos_data()
        auto_fit()

        # Salvando workbook(pasta de trabalho)
        wb.save(workbook)
        logging.info('Formatação concluída.')
    except PermissionError:
        logging.error('Feche o arquivo Excel antes de continuar.')
    except FileNotFoundError:
        logging.error('Arquivo não encontrado.')
    except KeyError:
        logging.error('Sheet não encontrada.')
    except Exception as e:
        logging.exception(f'Erro inesperado: {e}')

# Linha de montagem de uma fatura: lê o PDF, monta a tabela e confere o total.
# Entra: parser do banco, caminho do PDF, nome da aba. Sai: (df, feedback).
def processar(parser, arquivo, nome):
    dados, total = parser(arquivo)
    df = preparar_df(dados)
    soma = df['Valor'].sum()
    feedback = conferir(soma, total, nome)

    return df, feedback

# =================== CHAMADAS ====================

faturas = [] # Junta (df, feedback, nome) de cada banco — uma tupla por aba do workbook

# Identificação dos arquivos, leitura dos parsers, criação de df, exportação e manipulação de arquivo xlsx
for caminho_pdf in pasta_faturas.glob('*.pdf'):
    parser, nome = identificador(caminho_pdf)
    if parser is None:
        logging.warning(f'Banco não reconhecido: {caminho_pdf.name}')
        continue
    df, feedback = processar(parser, caminho_pdf, nome)
    faturas.append((df, feedback, nome))

# Exportar
exportar_xlsx(faturas, caminho_saida)

# Formatar
manipular_xlsx(caminho_saida)

# Fecha o cronômetro e loga o tempo total — referência pra comparar quando for otimizando
fim = perf_counter()
logging.info(f'Tempo total de execução: {fim - inicio:.2f}s')
